from openpyxl import load_workbook
from docx import Document


#1 Carrega o arquivo do Excel
caminho_excel = "fornecedores.xlsx" #certifique-se que o caminho esta correto

workbook = load_workbook(caminho_excel)
sheet = workbook.active

# 2 Lê o contrato base (agora diretamente no word)
for row in sheet.iter_rows(min_row=2, values_only=true): #Ignora o cabeçalho(linha 1)